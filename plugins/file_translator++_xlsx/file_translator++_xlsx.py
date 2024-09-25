from GalTransl import LOGGER
from GalTransl.GTPlugin import GFilePlugin
import openpyxl
import re


class file_plugin(GFilePlugin):
    def gtp_init(self, plugin_conf: dict, project_conf: dict):
        """
        This method is called when the plugin is loaded. 在插件加载时被调用。
        :param plugin_conf: The settings for the plugin. 插件yaml中所有设置的dict。
        :param project_conf: The settings for the project. 项目yaml中common下设置的dict。
        """
        # 打印提示的方法，打印时请带上模块名，以便区分日志。
        self.pname = plugin_conf["Core"].get("Name", "")
        settings = plugin_conf["Settings"]
        LOGGER.debug(
            f"[{self.pname}] 当前配置：是否自动识别名称:{settings.get('是否自动识别名称', True)}")
        # 读取配置文件中的设置，并保存到变量中。
        self.是否自动识别名称 = settings.get("是否自动识别名称", True)
        self.名称识别拼接方案 = settings.get("名称识别拼接方案", "{name}\n{message}")
        self.名称识别正则表达式 = re.compile(
            settings.get("名称识别正则表达式", r"^(?P<name>.*?)(?P<message>「.*?」)$"), re.DOTALL)
        self.清除换行符 = settings.get("清除换行符", True)
        self.ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
        self.ANSI_ESCAPE = re.compile(r'\x1B(?:[@-Z\\-_]|\[[0-?]*[ -/]*[@-~])')

    def load_file(self, file_path: str) -> list:
        """
        This method is called to load a file.
        加载文件时被调用。
        :param file_path: The path of the file to load. 文件路径。
        :return: A list of objects with message and name(optional). 返回一个包含message和name(可空)的对象列表。
        """
        if not file_path.endswith(".xlsx"):
            # 检查不支持的文件类型并抛出TypeError
            raise TypeError("File type not supported.")
        return self.read_xlsx_to_json(file_path)

    def save_file(self, file_path: str, transl_json: list):
        """
        This method is called to save a file. 保存文件时被调用。
        :param file_path: The path of the file to save. 保存文件路径。
        :param transl_json: A list of objects same as the return of load_file(). load_file提供的json在翻译message和name后的结果。
        :return: None.
        """
        EXCEL_TITLE_LIST = ["Original Text", "Initial",
                            "Machine translation", "Better translation", "Best translation"]

        LOGGER.debug(f"[{self.pname}] Saving file {file_path}")

        original_file_path_01 = file_path.replace("gt_output", "gt_input")
        original_file_path_02 = file_path.replace("gt_output", "json_jp")
        # 读取原文件
        original_file_path = original_file_path_01 or original_file_path_02

        try:
            # 读取原文件的Original Text
            original_texts = []
            if original_file_path:
                original_workbook = openpyxl.load_workbook(original_file_path)
                original_sheet = original_workbook.active
                original_texts = [cell.value for cell in original_sheet['A']]
                # 检查并跳过标题行
                if original_texts and original_texts[0] == "Original Text":
                    original_texts = original_texts[1:]

            # 创建一个新的工作簿
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            # 写入标题行
            for i, title in enumerate(EXCEL_TITLE_LIST):
                sheet.cell(row=1, column=i+1, value=self.check_string(title))

            # 重组翻译结果
            for i, transl_obj in enumerate(transl_json):
                if self.是否自动识别名称 and transl_obj["name"] is not None and transl_obj["name"] != "":
                    name = self.check_string(transl_obj["name"])
                    message = self.check_string(transl_obj["message"])
                    translated_text = self.名称识别拼接方案.format(name=name, message=message)
                else:
                    translated_text = self.check_string(transl_obj["message"])

                # 写入 Original Text 栏
                sheet.cell(
                    row=i+2, column=1, value=self.check_string(original_texts[i] if i < len(original_texts) else translated_text))
                # 写入 Machine translation 栏
                sheet.cell(row=i+2, column=3, value=translated_text)

            workbook.save(file_path)
        except Exception as e:
            LOGGER.error(f"保存文件 {file_path} 时出错: {e}")
            raise e


    def check_string(self, value):
        """检查字符串编码、长度和换行符"""
        if value is None:
            return value
        # 转换为str字符串
        original_value = str(value)
        
        # 移除ANSI转义序列
        value_without_ansi = self.ANSI_ESCAPE.sub('', original_value)
        if value_without_ansi != original_value:
            LOGGER.debug(f"[{self.pname}] 已移除ANSI转义序列: {original_value} -> {value_without_ansi}")
        
        # 移除其他非法字符
        value_without_illegal = self.ILLEGAL_CHARACTERS_RE.sub('', value_without_ansi)
        if value_without_illegal != value_without_ansi:
            LOGGER.debug(f"[{self.pname}] 已移除非法字符: {value_without_ansi} -> {value_without_illegal}")
        
        # 字符串长度不能超过32,767个字符
        # 如果需要，进行截断
        if len(value_without_illegal) > 32767:
            truncated_value = value_without_illegal[:32767]
            LOGGER.debug(f"[{self.pname}] 字符串已截断: {len(value_without_illegal)} -> 32767")
        else:
            truncated_value = value_without_illegal
        
        return truncated_value

    def gtp_final(self):
        """
        This method is called after all translations are done.
        在所有文件翻译完成之后的动作，例如输出提示信息。
        """
        pass

    def read_xlsx_to_json(self, file_path: str) -> list:
        """
        读取xlsx文件，返回一个包含message和name(可空)的对象列表。
        :param file_path: xlsx文件路径。
        :return: 包含message和name(可空)的对象列表。
        """
        json_list = []

        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            sheet = workbook.active

            # 判断文件是否是空的
            if sheet.max_row < 2:
                LOGGER.warning(f"File {file_path} is empty.")
                return json_list

            # 获取第一列的所有内容
            first_column_values = [row[0].value for row in sheet.iter_rows(
                min_col=1, max_col=1, min_row=1)]

            # 检查并跳过标题行
            if first_column_values and first_column_values[0] == "Original Text":
                first_column_values = first_column_values[1:]

            for row in first_column_values:

                # 跳过空行
                if row is None or row == "":
                    json_list.append({"name": "", "message": ""})
                    continue

                row = str(row).strip()
                name, message = "", row

                if self.是否自动识别名称:
                    match = self.名称识别正则表达式.search(row)
                    if match:
                        name = match.group("name").strip()
                        message = match.group("message").strip()
                if self.清除换行符:
                    message = message.replace("\n", " ")
                    message = message.replace("\r", " ")
                json_list.append({"name": name, "message": message})

        except FileNotFoundError:
            LOGGER.error(f"File not found: {file_path}")
            raise
        except openpyxl.utils.exceptions.InvalidFileException:
            LOGGER.error(f"Invalid file format: {file_path}")
            raise
        except Exception as e:
            LOGGER.error(f"Error loading file {file_path}: {e}")
            raise

        return json_list
